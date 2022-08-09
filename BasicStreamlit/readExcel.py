from email.policy import default
import streamlit as st
import pandas as pd
import openpyxl

# st.subheader("Dataset")
# data_file = st.file_uploader("Upload excel",type=['xlsx'])
# process_btn = st.button("Process",on_click=readExcelFile)

# if data_file is not None:
#     file_details = {"Filename":data_file.name,"FileType":data_file.type,"FileSize":data_file.size}
#     st.write(file_details)
#     wb = openpyxl.load_workbook(data_file)
#     sheet_selector = st.selectbox("Select sheetname",wb.sheetnames)     
#     st.markdown(f"# Currently Selected {sheet_selector}")
#     df = pd.read_excel(data_file,sheet_name=sheet_selector)
#     st.write(df)

# Nesting widgets after a button is not a good idea because it immediately goes back to a False state (There is a demo of that behavior in this video tutorial: https://youtu.be/EnXJBsCIl_A).

st.header("Sheet view")
data_file = st.sidebar.file_uploader("Upload Excel file",type=['xlsx'])  

if data_file:
    file_details = {
        "Filename":data_file.name,
        "FileType":data_file.type,
        "FileSize":data_file.size}

    wb = openpyxl.load_workbook(data_file)

    ## Show Excel file
    st.sidebar.subheader("File details:")
    st.sidebar.json(file_details,expanded=False)
    st.sidebar.markdown("----")

    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Select sheet:",wb.sheetnames)     
    df = pd.read_excel(data_file,sheet_selector)
    st.markdown(f"### Currently Selected: `{sheet_selector}`")
    st.write(df)

    ## Do something after a button
    doLogic_btn = st.button("➕")
    if doLogic_btn:
        df2 = df.sum().transpose()
        st.write(df2)

        # Do something else after the previous button
        another_btn = st.checkbox("Another +") #- but this will make the button go back to _False_ so nothing will be shown
        if another_btn:
            df3 = df2.sum()
            st.write(df3)    

